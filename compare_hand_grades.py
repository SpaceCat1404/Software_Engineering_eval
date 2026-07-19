"""
Compares this system's SRS/Test Plan/SADS scores against the professor's
hand-graded marks in "SE Project consolidated Marks.xlsx", to sanity-check
whether the automated grader tracks real grading.

Rewritten (2026-07-19) to join through pipeline_out/teams.db (team_db.py)
instead of matching submission filenames against individual spreadsheet
rows -- the old approach only worked when the submitting student happened
to be a team's first spreadsheet row (the only row marks are populated on),
so most submissions matched nothing. teams.db already maps every submitted
file to its team (by filename, falling back to searching document content
for an SRN/name), so this script only needs to: pick one file per team
(mtime-latest, when a team has multiple), find that file's score.json by
its recorded source filename, and compare against the team's hand mark.

Run team_db.py build first (or after any submission-set change) to refresh
pipeline_out/teams.db before running this.

Usage: python compare_hand_grades.py [path/to/consolidated_marks.xlsx]
"""
import sys
import glob
import json
import sqlite3
import statistics
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
XLSX_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "SE Project consolidated Marks.xlsx"
DB_PATH = ROOT / "pipeline_out" / "teams.db"
BLOB_DIR = ROOT / "pipeline_out" / "blob_local" / "submissions"

# doc_type -> (teams.db mark column, excel header for hand-graded max, our rubric's max_total)
DOC_TYPES = {
    "srs": ("srs_mark", "SRS", 12),
    "test_plan": ("test_plan_mark", "Test Plan", 13),  # our rubric is out of 13; excel's hand-graded max is 8
    "sads": ("sad_mark", "SAD", 20),
}


def load_hand_max(path):
    """Row 2 of each sheet carries the max marks per column (e.g. 'SRS' -> 12)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    max_marks = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header, max_row = rows[0], rows[1]
        col = {h: i for i, h in enumerate(header) if h}
        for key in ("SRS", "Test Plan", "SAD"):
            if key in col and max_row[col[key]] is not None:
                max_marks[key] = max_row[col[key]]
    return max_marks


def load_score_index(doc_type):
    """source filename (as recorded inside the score.json, i.e. the real
    submitted filename) -> parsed score dict. Blob filenames don't always
    equal the source filename 1:1 (some old blobs predate the dedup-by-
    team_id fix), so index by the recorded 'source' field, not the path."""
    index = {}
    for path in sorted(glob.glob(str(BLOB_DIR / doc_type / "*_score.json"))):
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        source = data.get("source")
        if source:
            index[source] = data
    return index


def pick_submission(rows):
    """Given a team's submission rows (sqlite3.Row, from the `submissions`
    table) for one doc_type, pick the one to compare: mtime-latest file on
    disk when there's more than one (user decision, 2026-07-19 -- earlier
    dedup rules like filename-suffix ranking don't cover every real case
    seen in this submission set, e.g. differently-named resubmissions)."""
    if len(rows) == 1:
        return rows[0]
    scored = []
    for r in rows:
        p = Path(r["filepath"])
        mtime = p.stat().st_mtime if p.exists() else -1
        scored.append((mtime, r))
    scored.sort(key=lambda t: t[0], reverse=True)
    return scored[0][1]


def main():
    if not DB_PATH.exists():
        print(f"{DB_PATH} not found -- run `python team_db.py build` first.")
        sys.exit(1)
    if not XLSX_PATH.exists():
        print(f"Excel file not found: {XLSX_PATH}")
        sys.exit(1)

    hand_max = load_hand_max(XLSX_PATH)
    print(f"Hand-graded max marks (from spreadsheet row 2): {hand_max}\n")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    for doc_type, (mark_col, excel_col, our_max) in DOC_TYPES.items():
        doc_dir = BLOB_DIR / doc_type
        if not doc_dir.exists():
            print(f"=== {doc_type.upper()} === no local blob data found, skipping\n")
            continue

        score_index = load_score_index(doc_type)
        h_max = hand_max.get(excel_col)

        sub_rows = conn.execute(
            "SELECT s.*, t.team_name, t.section, t.excel_team_id, t.{mc} AS hand_score "
            "FROM submissions s JOIN teams t ON s.team_pk = t.id "
            "WHERE s.doc_type = ? AND s.team_pk IS NOT NULL".format(mc=mark_col),
            (doc_type,),
        ).fetchall()

        by_team = {}
        for r in sub_rows:
            by_team.setdefault(r["team_pk"], []).append(r)

        rows = []
        no_score_found = []
        no_hand_mark = []

        for team_pk, team_subs in by_team.items():
            chosen = pick_submission(team_subs)
            tid = chosen["excel_team_id"]
            label = chosen["team_name"] or f"team #{int(tid) if tid is not None else '?'}"

            if chosen["hand_score"] is None:
                no_hand_mark.append((label, chosen["filename"]))
                continue

            entry = score_index.get(chosen["filename"])
            if entry is None or entry.get("total_score") is None:
                no_score_found.append((label, chosen["filename"]))
                continue

            hand_score = chosen["hand_score"]
            sys_max = entry.get("max_total") or our_max
            hand_pct = (hand_score / h_max * 100) if h_max else None
            sys_pct = entry["percentage"] if entry.get("percentage") is not None else (entry["total_score"] / sys_max * 100)
            sys_scaled_to_hand = entry["total_score"] * (h_max / sys_max) if h_max and sys_max else None

            rows.append({
                "name": label,
                "file": chosen["filename"],
                "picked_from": len(team_subs),
                "hand_score": hand_score,
                "hand_max": h_max,
                "hand_pct": hand_pct,
                "sys_score": entry["total_score"],
                "sys_max": sys_max,
                "sys_pct": sys_pct,
                "sys_scaled_to_hand": sys_scaled_to_hand,
                "diff_scaled": (hand_score - sys_scaled_to_hand) if sys_scaled_to_hand is not None else None,
                "diff_pct": (hand_pct - sys_pct) if hand_pct is not None else None,
            })

        print(f"=== {doc_type.upper()} ===  (our max={our_max}, hand-graded max={h_max})")
        print(f"{'Name':<32} {'Hand':>8} {'System':>10} {'Sys->Hand scale':>16} {'Diff':>7} {'Files':>6}")
        for r in sorted(rows, key=lambda r: (r["diff_pct"] is None, abs(r["diff_pct"] or 0)), reverse=True):
            multi = f"(1 of {r['picked_from']})" if r["picked_from"] > 1 else ""
            print(
                f"{r['name'][:32]:<32} "
                f"{r['hand_score']:>5.1f}/{int(r['hand_max']):<2} "
                f"{r['sys_score']:>6.1f}/{int(r['sys_max']):<3} "
                f"{r['sys_scaled_to_hand']:>13.2f}/{int(r['hand_max']):<2} "
                f"{r['diff_scaled']:>+6.2f} "
                f"{multi:>6}"
            )

        if rows:
            diffs = [r["diff_scaled"] for r in rows if r["diff_scaled"] is not None]
            mean_abs = statistics.mean(abs(d) for d in diffs)
            print(f"\nCompared: {len(rows)} teams   No score.json found: {len(no_score_found)}   No hand mark in spreadsheet: {len(no_hand_mark)}")
            print(f"Mean diff (hand - system, on hand's {h_max}-mark scale): {statistics.mean(diffs):+.2f}")
            print(f"Mean absolute diff: {mean_abs:.2f}  ({mean_abs / h_max * 100:.1f}% of {h_max}-mark scale)")
            print(f"Std dev of diff: {statistics.pstdev(diffs):.2f}  ({statistics.pstdev(diffs) / h_max * 100:.1f}% of {h_max}-mark scale)")
        else:
            print("No matches found.")

        if no_score_found:
            print(f"\nTeams matched but no score.json found for the picked file ({len(no_score_found)}):")
            for label, fname in no_score_found:
                print(f"  - {label}: {fname}")
        if no_hand_mark:
            print(f"\nTeams matched but spreadsheet has no {excel_col} mark ({len(no_hand_mark)}):")
            for label, fname in no_hand_mark:
                print(f"  - {label}: {fname}")

        print()

    conn.close()


if __name__ == "__main__":
    main()
