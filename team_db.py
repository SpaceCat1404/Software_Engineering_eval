"""
Builds a SQLite database that tracks TEAMS (as graded by the professor in
the consolidated marks spreadsheet) and maps each submitted file (SRS /
Test Plan / SADS) to the team it belongs to.

Why this exists: the spreadsheet groups students into teams (one row per
member, only the first member's row carries the team's marks), but only
ONE team member actually submits each document. compare_hand_grades.py
previously matched submissions to individual spreadsheet ROWS by parsed
filename name -- which is wrong, because (a) a team's marks only live on
its first row so a submission from a non-first member matched nothing,
and (b) it can't handle a case where the filename doesn't clearly name a
student who is in the spreadsheet at all.

Matching strategy per submission file:
  1. Parse a candidate name out of the filename (LMS-appended
     " - <NAME> PESU..." suffix) and fuzzy-match it against every member
     name of every team. Best match above a threshold wins.
  2. If that fails, open the actual file (PDF/docx, via srs_ingest.ingest)
     and search its extracted text for any team's member names or SRNs
     appearing literally in the document (e.g. a cover page). This is the
     "look into the PDF" fallback the filename-only approach can't do.
  3. If still nothing, leave it unmatched for manual review rather than
     guessing.

Usage:
  python team_db.py build     -- (re)build pipeline_out/teams.db from the
                                  spreadsheet + student_submissions/
  python team_db.py report    -- print match/coverage stats from the
                                  existing db without rebuilding
"""
import sys
import re
import sqlite3
import difflib
from pathlib import Path

import openpyxl

import srs_ingest

ROOT = Path(__file__).parent
XLSX_PATH = ROOT / "SE Project consolidated Marks.xlsx"
DB_PATH = ROOT / "pipeline_out" / "teams.db"

SUBMISSION_DIRS = {
    "srs": ROOT / "student_submissions" / "SRS_responses",
    "test_plan": ROOT / "student_submissions" / "test_plan_responses",
    "sads": ROOT / "student_submissions" / "SAD_spec_responses",
}

FILENAME_MATCH_THRESHOLD = 0.6
CONTENT_NAME_MATCH_THRESHOLD = 0.75


# ---------------------------------------------------------------- spreadsheet

def normalize_tokens(name):
    if not name:
        return frozenset()
    return frozenset(t for t in re.split(r"[^A-Za-z]+", name.upper()) if len(t) > 1)


def load_teams_from_excel(path=XLSX_PATH):
    """Returns a list of team dicts: {section, excel_team_id, team_name,
    project_id, project_name, srs_mark, test_plan_mark, sad_mark,
    test_cases_mark, impl_mark, overall_total, members: [{name, srn}]}.
    Forward-fills team-level fields across a team's member rows -- the
    spreadsheet only populates them on a team's first row, not via merged
    cells, so plain cell reads give None for rows 2-4 of each team."""
    wb = openpyxl.load_workbook(path, data_only=True)
    teams = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        col = {h: i for i, h in enumerate(header) if h}
        project_col = col.get("Project name") or col.get("Project Title")
        desc_col = col.get("Project description") or col.get("Project Description")

        current = None
        for row in rows[2:]:
            member_name = row[col["Team members"]]
            if member_name is None:
                continue
            member_name = str(member_name).strip()

            if row[col["Team ID"]] is not None:
                if current:
                    teams.append(current)
                current = {
                    "section": sheet_name,
                    "excel_team_id": row[col["Team ID"]],
                    "team_name": row[col.get("Team name")],
                    "project_id": row[col.get("Project ID")],
                    "project_name": row[project_col] if project_col is not None else None,
                    "srs_mark": row[col.get("SRS")],
                    "test_plan_mark": row[col.get("Test Plan")],
                    "sad_mark": row[col.get("SAD")],
                    "test_cases_mark": row[col.get("Test Cases")],
                    "impl_mark": row[col.get("Implementation total + demo")],
                    "overall_total": row[col.get("Overall total")],
                    "members": [],
                }
            if current is None:
                continue
            current["members"].append({
                "name": member_name,
                "srn": row[col.get("SRN")],
            })
        if current:
            teams.append(current)
    return teams


# ---------------------------------------------------------------- filenames

def extract_name_from_filename(filename):
    matches = list(re.finditer(r"PESU(?:-EC)?", filename))
    if not matches:
        return None
    prefix = filename[:matches[-1].start()]
    pos = prefix.rfind(" - ")
    if pos == -1:
        return None
    return prefix[pos + 3:].strip()


def name_similarity(a_tokens, b_tokens, a_str, b_str):
    if not a_tokens or not b_tokens:
        return 0.0
    overlap = len(a_tokens & b_tokens) / len(a_tokens | b_tokens)
    ratio = difflib.SequenceMatcher(None, a_str, b_str).ratio()
    return max(overlap, ratio)


def match_by_filename(filename, teams):
    candidate = extract_name_from_filename(filename)
    if not candidate:
        return None, None, 0.0
    cand_tokens = normalize_tokens(candidate)
    best_team, best_member, best_score = None, None, 0.0
    for team in teams:
        for member in team["members"]:
            score = name_similarity(cand_tokens, normalize_tokens(member["name"]), candidate, member["name"])
            if score > best_score:
                best_team, best_member, best_score = team, member["name"], score
    if best_score >= FILENAME_MATCH_THRESHOLD:
        return best_team, best_member, best_score
    return None, None, best_score


# ---------------------------------------------------------------- content fallback

def match_by_content(filepath, teams):
    """Ingests the actual file and searches its extracted text for a
    literal SRN (strongest signal -- unique per student) or a team
    member's full name. Used only when the filename didn't yield a
    confident match."""
    try:
        ingested = srs_ingest.ingest(filepath)
    except Exception as e:
        return None, None, 0.0, f"ingest failed: {e}"

    text = ingested.get("raw_text", "")
    text_upper = text.upper()

    for team in teams:
        for member in team["members"]:
            srn = member.get("srn")
            if srn and str(srn).upper() in text_upper:
                return team, member["name"], 1.0, "srn found in document text"

    best_team, best_member, best_score = None, None, 0.0
    for team in teams:
        for member in team["members"]:
            name = member["name"]
            if not name or len(name.strip()) < 4:
                continue
            if name.upper().strip() in text_upper:
                score = 0.95
            else:
                # fuzzy line-by-line scan is expensive; approximate via
                # difflib against the whole text is unreliable for a short
                # name in a long doc, so only accept exact substring here.
                continue
            if score > best_score:
                best_team, best_member, best_score = team, name, score

    if best_score >= CONTENT_NAME_MATCH_THRESHOLD:
        return best_team, best_member, best_score, "member name found in document text"
    return None, None, best_score, "no SRN or member name found in document text"


# ---------------------------------------------------------------- db

SCHEMA = """
CREATE TABLE teams (
    id INTEGER PRIMARY KEY,
    section TEXT,
    excel_team_id REAL,
    team_name TEXT,
    project_id REAL,
    project_name TEXT,
    srs_mark REAL,
    test_plan_mark REAL,
    sad_mark REAL,
    test_cases_mark REAL,
    impl_mark REAL,
    overall_total REAL
);
CREATE TABLE team_members (
    team_pk INTEGER REFERENCES teams(id),
    name TEXT,
    srn TEXT
);
CREATE TABLE submissions (
    doc_type TEXT,
    filename TEXT,
    filepath TEXT,
    team_pk INTEGER REFERENCES teams(id),
    matched_member TEXT,
    match_method TEXT,
    confidence REAL,
    note TEXT
);
"""


def build_db(db_path=DB_PATH, xlsx_path=XLSX_PATH, submission_dirs=None):
    submission_dirs = submission_dirs or SUBMISSION_DIRS
    teams = load_teams_from_excel(xlsx_path)

    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        db_path.unlink()
    conn = sqlite3.connect(db_path)
    conn.executescript(SCHEMA)

    team_pk_of = {}
    for i, team in enumerate(teams):
        cur = conn.execute(
            "INSERT INTO teams (section, excel_team_id, team_name, project_id, project_name, "
            "srs_mark, test_plan_mark, sad_mark, test_cases_mark, impl_mark, overall_total) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (team["section"], team["excel_team_id"], team["team_name"], team["project_id"],
             team["project_name"], team["srs_mark"], team["test_plan_mark"], team["sad_mark"],
             team["test_cases_mark"], team["impl_mark"], team["overall_total"]),
        )
        pk = cur.lastrowid
        team_pk_of[id(team)] = pk
        for member in team["members"]:
            conn.execute("INSERT INTO team_members (team_pk, name, srn) VALUES (?,?,?)",
                         (pk, member["name"], member["srn"]))
    conn.commit()

    stats = {}
    for doc_type, folder in submission_dirs.items():
        if not folder.exists():
            continue
        files = sorted(p for p in folder.iterdir() if p.suffix.lower() in (".pdf", ".docx") and not p.name.startswith("~$"))
        matched = content_matched = unmatched = 0

        for filepath in files:
            team, member, score = match_by_filename(filepath.name, teams)
            method = "filename"
            note = None

            if team is None:
                team, member, score, note = match_by_content(filepath, teams)
                method = "content"

            if team is None:
                unmatched += 1
                team_pk = None
            else:
                matched += 1
                if method == "content":
                    content_matched += 1
                team_pk = team_pk_of[id(team)]

            conn.execute(
                "INSERT INTO submissions (doc_type, filename, filepath, team_pk, matched_member, "
                "match_method, confidence, note) VALUES (?,?,?,?,?,?,?,?)",
                (doc_type, filepath.name, str(filepath), team_pk, member, method if team else "none", score, note),
            )

        stats[doc_type] = {"total": len(files), "matched": matched, "content_fallback": content_matched, "unmatched": unmatched}

    conn.commit()
    conn.close()
    return teams, stats


def print_report(db_path=DB_PATH):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    n_teams = conn.execute("SELECT COUNT(*) FROM teams").fetchone()[0]
    print(f"Teams in spreadsheet: {n_teams}\n")

    for doc_type in ("srs", "test_plan", "sads"):
        rows = conn.execute(
            "SELECT s.*, t.team_name, t.section, t.excel_team_id FROM submissions s LEFT JOIN teams t ON s.team_pk = t.id "
            "WHERE s.doc_type = ? ORDER BY s.team_pk IS NULL, t.section, t.team_name", (doc_type,)
        ).fetchall()
        if not rows:
            continue

        distinct_teams = {r["team_pk"] for r in rows if r["team_pk"] is not None}
        multi = {}
        for r in rows:
            if r["team_pk"] is not None:
                multi.setdefault(r["team_pk"], []).append(r["filename"])

        print(f"=== {doc_type.upper()} ===")
        print(f"{len(rows)} files -> {len(distinct_teams)} distinct teams matched, "
              f"{sum(1 for r in rows if r['team_pk'] is None)} unmatched, "
              f"{sum(1 for r in rows if r['match_method']=='content')} needed content fallback")

        dupes = {k: v for k, v in multi.items() if len(v) > 1}
        if dupes:
            print(f"\nTeams with >1 submitted file for {doc_type} (needs a human pick, not auto-deduped):")
            for team_pk, files in dupes.items():
                t = conn.execute("SELECT team_name, section, excel_team_id FROM teams WHERE id=?", (team_pk,)).fetchone()
                label = t["team_name"] or f"team #{int(t['excel_team_id'])}"
                print(f"  - {label} ({t['section']}): {files}")

        unmatched_rows = [r for r in rows if r["team_pk"] is None]
        if unmatched_rows:
            print(f"\nUnmatched files ({len(unmatched_rows)}):")
            for r in unmatched_rows:
                print(f"  - {r['filename']}  [{r['note'] or 'no candidate name found'}]")

        low_conf = [r for r in rows if r["team_pk"] is not None and r["confidence"] < 0.85]
        if low_conf:
            print(f"\nLow-confidence matches worth eyeballing ({len(low_conf)}):")
            for r in low_conf:
                label = r["team_name"] or f"team #{int(r['excel_team_id'])}"
                print(f"  - {r['filename']} -> {label} via {r['matched_member']} "
                      f"({r['match_method']}, confidence {r['confidence']:.2f})")

        print()

    conn.close()


def main():
    cmd = sys.argv[1] if len(sys.argv) > 1 else "build"
    if cmd == "build":
        teams, stats = build_db()
        print(f"Built {DB_PATH} with {len(teams)} teams.\n")
        for doc_type, s in stats.items():
            print(f"{doc_type}: {s['total']} files, {s['matched']} matched "
                  f"({s['content_fallback']} via content fallback), {s['unmatched']} unmatched")
        print()
        print_report()
    elif cmd == "report":
        print_report()
    else:
        print(f"Unknown command: {cmd}. Use 'build' or 'report'.")
        sys.exit(1)


if __name__ == "__main__":
    main()
